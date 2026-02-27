from io import BytesIO

from django.http import HttpResponse
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from workbench.planning.reporting import project_planning
from workbench.tools.formats import hours


class PlanningXLSXDocument:
    def __init__(self):
        self.workbook = Workbook()
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

    def to_response(self, filename):
        """Convert to HTTP response for download"""
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _is_dark_color(self, hex_color):
        """Determine if a color is dark (for choosing text color)"""
        try:
            # Convert hex to RGB
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)

            # Calculate luminance
            luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
            return luminance < 0.5
        except Exception:
            return False

    def gantt_chart(self, project):  # noqa: C901
        """Create a Gantt chart for a project with planned work and milestones"""

        # Use the existing planning reporting infrastructure
        planning_data = project_planning(project)

        if not planning_data["projects_offers"]:
            return

        # Get service type colors
        service_types = {st["id"]: st for st in planning_data["service_types"]}

        # Create the main Gantt chart sheet
        ws = self.workbook.create_sheet("Gantt Chart")

        # Build headers
        headers = [
            _("Type"),
            _("Title"),
            _("User/Provider"),
            _("Service Type"),
            _("Hours"),
            _("Start"),
            _("End"),
            _("Status"),
        ]

        # Add week headers
        week_start_col = len(headers) + 1
        headers.extend(
            f"KW {week_data['week']}" for week_data in planning_data["weeks"]
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Process project data
        current_row = 2

        for project_data in planning_data["projects_offers"]:
            project_info = project_data["project"]

            # Add milestones first
            milestones = project_info.get("milestones")
            if milestones:
                for milestone in milestones:
                    # Write basic info
                    ws.cell(row=current_row, column=1, value=_("Milestone"))
                    ws.cell(row=current_row, column=2, value=milestone["title"])
                    ws.cell(row=current_row, column=3, value="")
                    ws.cell(row=current_row, column=4, value="")
                    ws.cell(
                        row=current_row,
                        column=5,
                        value=hours(milestone["hours"]) if milestone["hours"] else "",
                    )
                    ws.cell(row=current_row, column=6, value=milestone["dow"])
                    ws.cell(row=current_row, column=7, value=milestone["dow"])
                    ws.cell(row=current_row, column=8, value=_("Milestone"))

                    # Add milestone markers (diamonds) in week columns
                    for col_idx, week_flag in enumerate(milestone["graphical_weeks"]):
                        col = week_start_col + col_idx
                        if week_flag:
                            cell = ws.cell(row=current_row, column=col, value="◆")
                            cell.fill = PatternFill(
                                start_color="FFA500",
                                end_color="FFA500",
                                fill_type="solid",
                            )  # Orange
                            cell.alignment = Alignment(horizontal="center")

                    current_row += 1

            # Add planned work
            if project_data.get("offers"):
                for offer_data in project_data["offers"]:
                    offer_info = offer_data["offer"]

                    for work_data in offer_data["work_list"]:
                        work_info = work_data["work"]

                        # Determine status
                        if work_info["is_provisional"]:
                            status = _("Provisional")
                        elif offer_info.get("is_accepted"):
                            status = _("Accepted")
                        elif offer_info.get("is_declined"):
                            status = _("Declined")
                        else:
                            status = _("Open")

                        # Get service type info
                        service_type_id = work_info.get("service_type_id")
                        service_type_info = service_types.get(service_type_id, {})
                        service_type_title = service_type_info.get("title", "")
                        service_type_color = service_type_info.get(
                            "color", "#3498db"
                        )  # Default blue

                        # Remove # from color if present
                        service_type_color = service_type_color.removeprefix("#")

                        # Write basic info
                        ws.cell(row=current_row, column=1, value=_("Planned Work"))
                        ws.cell(row=current_row, column=2, value=work_info["title"])
                        ws.cell(row=current_row, column=3, value=work_info["user"])
                        ws.cell(row=current_row, column=4, value=service_type_title)
                        ws.cell(
                            row=current_row,
                            column=5,
                            value=hours(work_info["planned_hours"]),
                        )
                        ws.cell(row=current_row, column=6, value=work_info["range"])
                        ws.cell(row=current_row, column=7, value=work_info["range"])
                        ws.cell(row=current_row, column=8, value=status)

                        # Add colored bars for planned work weeks
                        for col_idx, week_hours in enumerate(
                            work_data["hours_per_week"]
                        ):
                            col = week_start_col + col_idx
                            if week_hours:
                                cell = ws.cell(
                                    row=current_row, column=col, value=hours(week_hours)
                                )
                                # Create colored background based on service type
                                cell.fill = PatternFill(
                                    start_color=service_type_color,
                                    end_color=service_type_color,
                                    fill_type="solid",
                                )
                                cell.alignment = Alignment(horizontal="center")
                                cell.font = Font(
                                    color="FFFFFF"
                                    if self._is_dark_color(service_type_color)
                                    else "000000"
                                )

                        current_row += 1

            # Add external work
            external_work = project_data.get("external_work")
            if external_work:
                for ext_work in external_work:
                    # Write basic info
                    ws.cell(row=current_row, column=1, value=_("External Work"))
                    ws.cell(row=current_row, column=2, value=ext_work["title"])
                    ws.cell(row=current_row, column=3, value=ext_work["provided_by"])
                    ws.cell(row=current_row, column=4, value="")
                    ws.cell(row=current_row, column=5, value="")
                    ws.cell(row=current_row, column=6, value=ext_work["range"])
                    ws.cell(row=current_row, column=7, value=ext_work["range"])
                    ws.cell(row=current_row, column=8, value=_("External"))

                    # Add external work markers
                    for col_idx, week_flag in enumerate(ext_work["by_week"]):
                        col = week_start_col + col_idx
                        if week_flag:
                            cell = ws.cell(row=current_row, column=col, value="●")
                            cell.fill = PatternFill(
                                start_color="808080",
                                end_color="808080",
                                fill_type="solid",
                            )  # Gray
                            cell.alignment = Alignment(horizontal="center")
                            cell.font = Font(color="FFFFFF")

                    current_row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add summary sheet
        self._add_summary_sheet(planning_data)

        # Add milestone details sheet
        self._add_milestone_details_sheet(planning_data)

        # Add capacity analysis sheet
        self._add_capacity_analysis_sheet(planning_data)

    def _add_summary_sheet(self, planning_data):
        """Add a summary sheet with project overview"""
        ws = self.workbook.create_sheet("Summary")

        # Calculate summary data
        total_planned_hours = sum(
            offer_data["offer"]["planned_hours"]
            for project_data in planning_data["projects_offers"]
            for offer_data in project_data.get("offers", [])
        )

        total_milestones = sum(
            len(project_data["project"].get("milestones") or [])
            for project_data in planning_data["projects_offers"]
        )

        total_external_work = sum(
            len(project_data.get("external_work") or [])
            for project_data in planning_data["projects_offers"]
        )

        # Write summary data
        row = 1

        # Project info
        if planning_data["projects_offers"]:
            proj = planning_data["projects_offers"][0]["project"]
            ws.cell(row=row, column=1, value=_("Project")).font = Font(bold=True)
            ws.cell(row=row, column=2, value=proj["title"])
            row += 1

            ws.cell(row=row, column=1, value=_("Status")).font = Font(bold=True)
            ws.cell(
                row=row, column=2, value=_("Closed") if proj["is_closed"] else _("Open")
            )
            row += 2

        # Totals
        ws.cell(row=row, column=1, value=_("Total Planned Hours")).font = Font(
            bold=True
        )
        ws.cell(row=row, column=2, value=hours(total_planned_hours))
        row += 1

        ws.cell(row=row, column=1, value=_("Total Milestones")).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_milestones)
        row += 1

        ws.cell(row=row, column=1, value=_("Total External Work Items")).font = Font(
            bold=True
        )
        ws.cell(row=row, column=2, value=total_external_work)

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _add_milestone_details_sheet(self, planning_data):
        """Add milestone details sheet if milestones exist"""
        milestones_exist = any(
            project_data["project"].get("milestones")
            for project_data in planning_data["projects_offers"]
        )

        if not milestones_exist:
            return

        ws = self.workbook.create_sheet("Milestone Details")

        # Headers
        headers = [_("Title"), _("Date"), _("Estimated Hours"), _("Phase Duration")]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Data
        row = 2
        for project_data in planning_data["projects_offers"]:
            milestones = project_data["project"].get("milestones")
            if milestones:
                for milestone in milestones:
                    ws.cell(row=row, column=1, value=milestone["title"])
                    ws.cell(row=row, column=2, value=milestone["dow"])
                    ws.cell(
                        row=row,
                        column=3,
                        value=hours(milestone["hours"]) if milestone["hours"] else "",
                    )
                    ws.cell(row=row, column=4, value=milestone.get("range", ""))
                    row += 1

        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _add_capacity_analysis_sheet(self, planning_data):
        """Add capacity analysis sheet if capacity data exists"""
        if not planning_data.get("capacity"):
            return

        ws = self.workbook.create_sheet("Capacity Analysis")

        # Headers
        headers = [_("User")] + [f"KW {w['week']}" for w in planning_data["weeks"]]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Total capacity row
        row = 2
        ws.cell(row=row, column=1, value=_("Total Capacity")).font = Font(bold=True)
        for col, capacity in enumerate(planning_data["capacity"]["total"], 2):
            ws.cell(row=row, column=col, value=hours(capacity))
        row += 1

        # Individual user capacity
        for user_data in planning_data["capacity"]["by_user"]:
            ws.cell(row=row, column=1, value=user_data["user"]["name"])
            for col, capacity in enumerate(user_data["capacity"], 2):
                ws.cell(row=row, column=col, value=hours(capacity))
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
